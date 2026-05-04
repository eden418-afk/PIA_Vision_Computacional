import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent

RESULTADOS_DIR = BASE_DIR / "resultados"
CSV_PATH = RESULTADOS_DIR / "cuadro_experimental.csv"
XLSX_PATH = RESULTADOS_DIR / "cuadro_experimental.xlsx"


def cargar_csv():
    filas = []

    with open(CSV_PATH, mode="r", encoding="utf-8") as file:
        reader = csv.DictReader(file)

        for row in reader:
            filas.append(row)

    return filas


def to_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return value


def promedio(valores):
    if not valores:
        return 0
    return sum(valores) / len(valores)


def contar_por_columna(filas, columna):
    conteo = {}

    for fila in filas:
        valor = fila.get(columna, "sin_dato")
        conteo[valor] = conteo.get(valor, 0) + 1

    return conteo


def aplicar_estilo_tabla(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def ajustar_columnas(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 32)


def crear_hoja_datos(wb, filas):
    ws = wb.active
    ws.title = "Datos experimentales"

    headers = list(filas[0].keys())
    ws.append(headers)

    for fila in filas:
        ws.append([to_float(fila[h]) for h in headers])

    aplicar_estilo_tabla(ws)
    ajustar_columnas(ws)

    return ws


def crear_hoja_resumen(wb, filas):
    ws = wb.create_sheet("Resumen estadístico")

    errores_centroide = [float(f["error_centroide_mm"]) for f in filas]
    errores_angulares = [float(f["error_angular_grados"]) for f in filas]
    areas = [float(f["area_px2"]) for f in filas]

    resumen = [
        ["Variable", "Valor"],
        ["Número de imágenes procesadas", len(filas)],

        ["Error promedio de centroide (mm)", round(promedio(errores_centroide), 4)],
        ["Error máximo de centroide (mm)", round(max(errores_centroide), 4)],
        ["Error mínimo de centroide (mm)", round(min(errores_centroide), 4)],

        ["Error promedio angular (grados)", round(promedio(errores_angulares), 4)],
        ["Error máximo angular (grados)", round(max(errores_angulares), 4)],
        ["Error mínimo angular (grados)", round(min(errores_angulares), 4)],

        ["Área promedio detectada (px²)", round(promedio(areas), 4)],
        ["Área máxima detectada (px²)", round(max(areas), 4)],
        ["Área mínima detectada (px²)", round(min(areas), 4)],
    ]

    for row in resumen:
        ws.append(row)

    aplicar_estilo_tabla(ws)
    ajustar_columnas(ws)

    return ws


def crear_hoja_discriminacion(wb, filas):
    ws = wb.create_sheet("Discriminación geométrica")

    headers = [
        "imagen",
        "figura_cad",
        "figura_clasificada_python",
        "numero_agujeros",
        "vertices_aprox",
        "aspect_ratio",
        "extent",
        "solidity",
        "circularity",
        "area_px2",
        "perimetro_px",
    ]

    ws.append(headers)

    for fila in filas:
        ws.append([
            fila.get("imagen"),
            fila.get("figura"),
            fila.get("figura_clasificada_python"),
            to_float(fila.get("numero_agujeros")),
            to_float(fila.get("vertices_aprox")),
            to_float(fila.get("aspect_ratio")),
            to_float(fila.get("extent")),
            to_float(fila.get("solidity")),
            to_float(fila.get("circularity")),
            to_float(fila.get("area_px2")),
            to_float(fila.get("perimetro_px")),
        ])

    aplicar_estilo_tabla(ws)
    ajustar_columnas(ws)

    return ws


def crear_hoja_conteos(wb, filas):
    ws = wb.create_sheet("Conteos clasificación")

    conteo_cad = contar_por_columna(filas, "figura")
    conteo_python = contar_por_columna(filas, "figura_clasificada_python")

    ws.append(["Tipo", "Categoría", "Cantidad"])

    for categoria, cantidad in conteo_cad.items():
        ws.append(["CAD", categoria, cantidad])

    for categoria, cantidad in conteo_python.items():
        ws.append(["Python", categoria, cantidad])

    aplicar_estilo_tabla(ws)
    ajustar_columnas(ws)

    return ws


def crear_hoja_configuracion(wb):
    ws = wb.create_sheet("Configuración")

    datos = [
        ["Parámetro", "Valor", "Descripción"],
        ["Ancho de imagen", "1280 px", "Resolución horizontal de las imágenes sintéticas"],
        ["Alto de imagen", "720 px", "Resolución vertical de las imágenes sintéticas"],
        ["Ancho de mesa", "320 mm", "Dimensión usada para convertir píxeles a milímetros"],
        ["Alto de mesa", "180 mm", "Dimensión usada para convertir píxeles a milímetros"],
        ["Referencia CAD", "FreeCAD", "Centroide e inclinación obtenidos desde modelos CAD"],
        ["Método de segmentación", "Binarización", "Separación de pieza oscura contra fondo blanco"],
        ["Método de bordes", "Canny", "Usado para visualización de contornos"],
        ["Método de centroide", "Momentos de imagen", "Centroide calculado desde máscara binaria"],
        ["Método de orientación", "Momentos centrales", "Eje principal estimado por distribución de masa"],
        ["Discriminación geométrica", "Reglas por propiedades", "Uso de área, perímetro, agujeros, vértices, solidez y circularidad"],
        ["Transformación al robot", "Matriz homogénea", "Conversión de coordenadas locales de mesa al sistema del robot"],
    ]

    for row in datos:
        ws.append(row)

    aplicar_estilo_tabla(ws)
    ajustar_columnas(ws)

    return ws


def agregar_graficas_datos(ws_datos):
    headers = [cell.value for cell in ws_datos[1]]

    col_imagen = headers.index("imagen") + 1
    col_error_centroide = headers.index("error_centroide_mm") + 1
    col_error_angular = headers.index("error_angular_grados") + 1

    max_row = ws_datos.max_row

    cats = Reference(ws_datos, min_col=col_imagen, min_row=2, max_row=max_row)

    chart1 = BarChart()
    chart1.title = "Error de centroide CAD vs Python"
    chart1.y_axis.title = "Error (mm)"
    chart1.x_axis.title = "Imagen"

    data1 = Reference(ws_datos, min_col=col_error_centroide, min_row=1, max_row=max_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 8
    chart1.width = 18

    ws_datos.add_chart(chart1, "AD2")

    chart2 = LineChart()
    chart2.title = "Error angular CAD vs Python"
    chart2.y_axis.title = "Error (grados)"
    chart2.x_axis.title = "Imagen"

    data2 = Reference(ws_datos, min_col=col_error_angular, min_row=1, max_row=max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 8
    chart2.width = 18

    ws_datos.add_chart(chart2, "AD18")


def agregar_grafica_conteos(ws_conteos):
    chart = BarChart()
    chart.title = "Conteo de figuras clasificadas"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Categoría"

    max_row = ws_conteos.max_row

    data = Reference(ws_conteos, min_col=3, min_row=1, max_row=max_row)
    cats = Reference(ws_conteos, min_col=2, min_row=2, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22

    ws_conteos.add_chart(chart, "E2")


def generar_excel():
    if not CSV_PATH.exists():
        raise FileNotFoundError(f"No existe el archivo: {CSV_PATH}")

    filas = cargar_csv()

    if not filas:
        raise ValueError("El CSV no contiene filas.")

    wb = Workbook()

    ws_datos = crear_hoja_datos(wb, filas)
    crear_hoja_resumen(wb, filas)
    crear_hoja_discriminacion(wb, filas)
    ws_conteos = crear_hoja_conteos(wb, filas)
    crear_hoja_configuracion(wb)

    agregar_graficas_datos(ws_datos)
    agregar_grafica_conteos(ws_conteos)

    wb.save(XLSX_PATH)

    print("Excel generado correctamente.")
    print(f"Archivo: {XLSX_PATH}")


if __name__ == "__main__":
    generar_excel()